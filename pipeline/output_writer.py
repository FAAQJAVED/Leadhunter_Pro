"""
pipeline.output_writer — CSV (UTF-8-BOM) + Excel (.xlsx) output writer.
"""

from __future__ import annotations

import csv
import logging
import os
from collections import Counter
from collections.abc import Iterable
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR
from pipeline.data_cleaner import CleanRecord

logger = logging.getLogger('lead_engine.output')

# ---------------------------------------------------------------------------
# Column definitions (order matters — matches CSV/xlsx column order)
# ---------------------------------------------------------------------------
COLUMNS = [
    'score',
    'company_name',
    'website_url',
    'domain',
    'search_query',
    'search_engine',
    'date_found',
    'flagged',
    'flag_reason',
]

COLUMN_HEADERS = {
    'score':        'Score',
    'company_name': 'Company Name',
    'website_url':  'Website URL',
    'domain':       'Domain',
    'search_query': 'Search Query',
    'search_engine': 'Search Engine',
    'date_found':   'Date Found',
    'flagged':      'Flagged',
    'flag_reason':  'Flag Reason',
}

_HEADER_FILL         = PatternFill('solid', fgColor='1F4E79')
_FLAG_FILL           = PatternFill('solid', fgColor='FFFF00')
_ALT_FILL            = PatternFill('solid', fgColor='EBF3FB')
_SUMMARY_HEADER_FILL = PatternFill('solid', fgColor='2E75B6')


class OutputWriter:
    """
    Writes accumulated CleanRecord objects to:
      - outputs/leads_YYYY-MM-DD_<engines>.csv  (UTF-8-BOM)
      - outputs/leads_YYYY-MM-DD_<engines>.xlsx

    Parameters
    ----------
    engines : List of engine names used in this run (e.g. ['mojeek', 'duckduckgo']).
              If provided, the engine name(s) are included in the output filename
              instead of an HH-MM timestamp so files are immediately recognisable.
              E.g. leads_2026-05-03_mojeek.csv or leads_2026-05-03_mojeek-ddg.csv.
              If None or empty, falls back to the old HH-MM timestamp format.
    """

    def __init__(self, engines: list[str] | None = None) -> None:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        date_str = datetime.now().strftime('%Y-%m-%d')
        if engines:
            # Shorten well-known engine names for compact filenames
            _SHORT = {"duckduckgo": "ddg", "mojeek": "mojeek",
                      "yahoo": "yahoo", "bing": "bing"}
            short = "-".join(_SHORT.get(e.lower(), e.lower()) for e in engines)
            stem = f'leads_{date_str}_{short}'
        else:
            ts   = datetime.now().strftime('%H-%M')
            stem = f'leads_{date_str}_{ts}'
        self.csv_path  = Path(OUTPUT_DIR) / f'{stem}.csv'
        self.xlsx_path = Path(OUTPUT_DIR) / f'{stem}.xlsx'
        self._records: list[CleanRecord] = []

    def add(self, record: CleanRecord) -> None:
        self._records.append(record)

    def add_many(self, records: Iterable[CleanRecord]) -> None:
        for r in records:
            self._records.append(r)

    def count(self) -> int:
        return len(self._records)

    def write_csv(self) -> str:
        """Write UTF-8-BOM CSV (opens cleanly in Excel). Returns path."""
        with open(self.csv_path, 'w', encoding='utf-8-sig', newline='') as fh:
            writer = csv.DictWriter(fh, fieldnames=COLUMNS)
            writer.writerow({col: COLUMN_HEADERS[col] for col in COLUMNS})
            for rec in self._records:
                writer.writerow({col: rec.to_dict()[col] for col in COLUMNS})
        logger.info("CSV written: %s (%d rows)", self.csv_path, len(self._records))
        return str(self.csv_path)

    def write_xlsx(self) -> str:
        """Write formatted .xlsx with data sheet + summary sheet."""
        wb = openpyxl.Workbook()
        self._write_data_sheet(wb)
        self._write_summary_sheet(wb)
        wb.save(self.xlsx_path)
        logger.info("Excel written: %s", self.xlsx_path)
        return str(self.xlsx_path)

    def write_all(self) -> tuple[str, str]:
        """Sort records by score desc, then write both files. Returns (csv_path, xlsx_path)."""
        self._records.sort(key=lambda r: r.score, reverse=True)
        return self.write_csv(), self.write_xlsx()

    def _write_data_sheet(self, wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws.title = 'Leads'
        ws.freeze_panes = 'A2'

        header_font  = Font(bold=True, color='FFFFFF', size=11)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border  = Border(
            bottom=Side(style='thin', color='AAAAAA'),
            right=Side(style='thin', color='AAAAAA'),
        )

        for col_idx, col_key in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=COLUMN_HEADERS[col_key])
            cell.font      = header_font
            cell.fill      = _HEADER_FILL
            cell.alignment = header_align
            cell.border    = thin_border

        ws.row_dimensions[1].height = 22

        data_font    = Font(size=10)
        data_align   = Alignment(vertical='top', wrap_text=False)
        flagged_font = Font(size=10, bold=True, color='7F0000')

        for row_idx, rec in enumerate(self._records, start=2):
            d          = rec.to_dict()
            is_flagged = bool(d['flagged'])
            is_alt     = (row_idx % 2 == 0)

            for col_idx, col_key in enumerate(COLUMNS, start=1):
                val  = d[col_key]
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

                if is_flagged:
                    cell.fill = _FLAG_FILL
                    cell.font = flagged_font
                elif is_alt:
                    cell.fill = _ALT_FILL
                    cell.font = data_font
                else:
                    cell.font = data_font

                cell.alignment = data_align

                if col_key == 'website_url' and isinstance(val, str) and val.startswith('http'):
                    cell.hyperlink = val
                    cell.font = (
                        Font(size=10, color='0563C1', underline='single', bold=is_flagged)
                        if not is_flagged else
                        Font(size=10, color='7F0000', underline='single', bold=True)
                    )

        col_widths = {
            'score': 8, 'company_name': 35, 'website_url': 50, 'domain': 30,
            'search_query': 40, 'search_engine': 14, 'date_found': 22,
            'flagged': 9, 'flag_reason': 20,
        }
        for col_idx, col_key in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = \
                col_widths.get(col_key, 20)

        ws.auto_filter.ref = (
            f'A1:{get_column_letter(len(COLUMNS))}{len(self._records) + 1}'
        )

    def _write_summary_sheet(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet('Summary')
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        h_font   = Font(bold=True, color='FFFFFF', size=12)
        h_fill   = _SUMMARY_HEADER_FILL
        v_font   = Font(size=11)
        lbl_font = Font(bold=True, size=11)

        def _hdr(row: int, text: str) -> None:
            cell = ws.cell(row=row, column=1, value=text)
            cell.font = h_font
            cell.fill = h_fill
            ws.merge_cells(f'A{row}:B{row}')
            cell.alignment = Alignment(horizontal='left')

        def _row(row: int, label: str, value) -> None:
            a = ws.cell(row=row, column=1, value=label)
            b = ws.cell(row=row, column=2, value=value)
            a.font = lbl_font
            b.font = v_font
            b.alignment = Alignment(horizontal='right')

        total   = len(self._records)
        flagged = sum(1 for r in self._records if r.flagged)
        unique  = len({r.domain for r in self._records})
        by_eng  = Counter(r.search_engine for r in self._records)

        row = 1
        _hdr(row, '📊 SCRAPE SUMMARY')
        row += 1
        _row(row, 'Total Records', total)
        row += 1
        _row(row, 'Unique Domains', unique)
        row += 1
        _row(row, 'Flagged Records', flagged)
        row += 1
        _row(row, 'Clean Records', total - flagged)
        row += 1
        _row(row, 'Generated', datetime.now().strftime('%Y-%m-%d %H:%M'))
        row += 2

        _hdr(row, '🔍 Results by Engine')
        row += 1
        for engine, count in sorted(by_eng.items()):
            _row(row, engine.capitalize(), count)
            row += 1
        row += 1

        by_query = Counter(r.search_query for r in self._records)
        _hdr(row, '🔎 Top Queries (by records found)')
        row += 1
        for query, count in by_query.most_common(20):
            _row(row, query[:45], count)
            row += 1
